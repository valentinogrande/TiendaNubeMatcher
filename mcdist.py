#!/usr/bin/env python3

import difflib
import json
import math
import re
import signal
import unicodedata
from collections import defaultdict
from functools import lru_cache

import requests
from openpyxl import load_workbook

# ============================================================
# CONFIG
# ============================================================

def signal_handler(sig, frame):

    print()
    print(f"{YELLOW}Interrupted by user (CTRL+C){RESET}")
    exit()

signal.signal(signal.SIGINT, signal_handler)

ACCESS_TOKEN = "e809b222306b2f94fcc0431c38147e2ca1dc9b5bad76e88c5bdb8a5f4ecfed04"

STORE_ID = "5975004"
USER_ID = "6191153"

BASE_URL = "https://cirrus.tiendanube.com"

LOCAL_FILE = "products.xlsx"

AUTO_CONFIRM_SAFE = False

MIN_SCORE = 40

# ============================================================
# HEADERS
# ============================================================

HEADERS = {
    "x-admin-platform": "web",
    "x-admin-front-version": "2026.4.30",
    "x-admin-build-version": "2f2aead",
    "x-access-token": ACCESS_TOKEN,
    "x-store-id": STORE_ID,
    "x-user-id": USER_ID,
    "x-authentication-factor-code": "no-defined",
    "x-store-country": "AR",
    "Content-Type": "application/json",
    "User-Agent": (
        "Mozilla/5.0 (X11; Linux x86_64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/146.0.0.0 Safari/537.36"
    ),
}

# ============================================================
# COLORS
# ============================================================

RED = "\033[91m"
GREEN = "\033[92m"
YELLOW = "\033[93m"
BLUE = "\033[94m"
MAGENTA = "\033[95m"
CYAN = "\033[96m"
RESET = "\033[0m"

# ============================================================
# STOPWORDS
# ============================================================

STOPWORDS = {
    "x",
    "kg",
    "gr",
    "g",
    "de",
    "la",
    "el",
    "los",
    "las",
    "y",
    "con",
    "por",
    "unidad",
    "unidades",
    "pack",
    "u",
    "uni",
}

# ============================================================
# BRANDS
# ============================================================


CRITICAL_BRANDS = {
    "paladini",
    "fela",
    "sadia",
    "la martina",
    "vasconia",
    "cremac",
    "pacama",
    "albor",
    "san cayetano",
    "rosamonte",
    "andresito",
    "playadito",
    "cbse",
    "terrabusi",
    "lucchetti",
    "galletti",
    "mocovi",
    "cabaña espinillo",
    "punta del agua",
    "lealtad",
    "molto",
    "arcor",
    "baggio",
    "mora",
    "milkey",
    "oreo",
    "aguila",
    "erevan",
    "veronica",
    "bon o bon",
    "pronatte",
    "regio",
}

SECTION_HEADERS = {
    "cucharitas",
    "frutas secas",
    "leche industrial",
    "termicos",
    "mantecol",
    "chocolate",
    "chocolates aguila",
}


# ============================================================
# SYNONYMS
# ============================================================

SYNONYMS = {
    "muzza": "mozzarella",
    "mozza": "mozzarella",
    "jyq": "jamon queso",
    "paleta cocida": "paleta",
    "regianito": "reggiano",
}

# ============================================================
# IMPORTANT TOKENS
# ============================================================

IMPORTANT_TOKENS = {
    "reggiano": 40,
    "sardo": 35,
    "mortadela": 35,
    "jamon": 35,
    "paleta": 35,
    "salchicha": 35,
    "mozzarella": 40,
    "cheddar": 40,
    "romano": 40,
    "fontina": 40,
    "holanda": 40,
    "cremoso": 30,
    "salut": 30,
}

# ============================================================
# GENERIC WORDS
# ============================================================

GENERIC_WORDS = {
    "barra",
    "cremoso",
    "light",
}

# ============================================================
# FIAMBRE TYPES
# ============================================================

FIAMBRE_TYPES = {
    "jamon",
    "paleta",
    "mortadela",
    "salchicha",
    "milan",
}

# ============================================================
# CATEGORY KEYWORDS
# ============================================================

CATEGORY_KEYWORDS = {
    "queso": [
        "cremoso",
        "mozzarella",
        "barra",
        "sardo",
        "fontina",
        "holanda",
        "danbo",
        "reggiano",
        "romano",
        "por salut",
        "cheddar",
    ],
    "limpieza": [
        "detergente",
        "lavandina",
        "jabon",
    ],
    "yerba": [
        "yerba",
        "mate",
        "rosamonte",
        "playadito",
        "cbse",
        "andresito",
    ],
    "fideos": [
        "fideos",
        "spaghetti",
        "tirabuzon",
        "mostachol",
    ],
    "arroz": [
        "arroz",
    ],
    "tomate": [
        "tomate",
        "pure",
        "salsa",
    ],
    "aceitunas": [
        "aceituna",
        "aceitunas",
    ],
    "fiambre": [
        "mortadela",
        "jamon",
        "paleta",
        "salchicha",
        "milan",
    ],
    "lacteo": [
        "leche",
        "crema",
    ],
}

# ============================================================
# INVALID CATEGORY COMBINATIONS
# ============================================================

INVALID_CATEGORY_COMBINATIONS = {
    ("queso", "limpieza"),
    ("yerba", "fiambre"),
    ("aceitunas", "fideos"),
    ("tomate", "limpieza"),
}

# ============================================================
# NORMALIZE
# ============================================================


@lru_cache(maxsize=50000)
def normalize(text):

    text = unicodedata.normalize("NFKD", text)
    text = text.encode("ascii", "ignore").decode("ascii")

    text = text.lower()

    for k, v in SYNONYMS.items():
        text = text.replace(k, v)

    text = re.sub(r"[^\w\s]", " ", text)

    text = re.sub(r"\s+", " ", text)

    return text.strip()


# ============================================================
# TOKENIZE
# ============================================================


def tokenize(text):

    text = normalize(text)

    tokens = []

    for t in text.split():

        if t in STOPWORDS:
            continue

        t = re.sub(r"^(\d+)(?:kg|g|u|uni|unid|lt|l)$", r"\1", t, flags=re.I)

        if len(t) > 4 and t.endswith("es"):
            t = t[:-2]
        elif len(t) > 3 and t.endswith("s"):
            t = t[:-1]

        tokens.append(t)

    return tokens


# ============================================================
# SIMILARITY
# ============================================================


def similarity(a, b):

    return difflib.SequenceMatcher(None, a, b).ratio()


# ============================================================
# CATEGORY
# ============================================================


def detect_category(text):

    t = normalize(text)

    for category, kws in CATEGORY_KEYWORDS.items():

        for kw in kws:

            if kw in t:
                return category

    return None


# ============================================================
# BRAND
# ============================================================


def extract_brand(text):

    t = normalize(text)

    found = []

    for brand in CRITICAL_BRANDS:

        if normalize(brand) in t:
            found.append(brand)

    if not found:
        return None

    found.sort(key=len, reverse=True)

    return found[0]


# ============================================================
# STRUCTURED UNITS
# ============================================================


def extract_structured_units(text):

    t = normalize(text)

    result = {
        "kg": None,
        "g": None,
        "units": None,
    }

    kg = re.findall(r"(\d+(?:\.\d+)?)\s*kg", t)

    if kg:
        result["kg"] = float(kg[0])

    g = re.findall(r"(\d+(?:\.\d+)?)\s*g", t)

    if g:
        result["g"] = float(g[0])

    units = re.findall(r"\bx\s*(\d+(?:\.\d+)?)", t)

    if units:
        result["units"] = int(units[0])

    return result


# ============================================================
# PRICE DIFF
# ============================================================


def percent_diff(a, b):

    if a <= 0 or b <= 0:
        return 999

    return abs(a - b) / max(a, b)


# ============================================================
# PARSE LOCAL PRODUCTS
# ============================================================

def parse_xlsx_local_products(path):

    products = []

    wb = load_workbook(path, data_only=True)

    for sheet_name in wb.sheetnames:

        ws = wb[sheet_name]

        current_brand = ""

        for row in ws.iter_rows(values_only=True):

            values = [v for v in row if v is not None]

            if not values:
                continue

            # =========================================================
            # HEADER DETECTION
            # =========================================================

            first = str(values[0]).strip()

            if (
                isinstance(values[0], str)
                and len(values) == 1
                and len(first.split()) <= 4
            ):

                current_brand = (
                    first.replace('"', "")
                    .replace(":", "")
                    .strip()
                    .lower()
                )

                continue

            # =========================================================
            # PRICE
            # =========================================================

            price = None

            for v in reversed(values):

                if isinstance(v, (int, float)) and v > 0:

                    # avoid weights interpreted as prices
                    if v < 100:
                        continue

                    price = int(v)
                    break

            if not price:
                continue

            # =========================================================
            # TEXT PARTS
            # =========================================================

            seen = set()

            text_parts = []

            for v in values:

                if isinstance(v, (int, float)):
                    continue

                s = str(v).strip()

                if not s:
                    continue

                norm = normalize(s)

                # remove duplicate columns
                if norm in seen:
                    continue

                seen.add(norm)

                text_parts.append(s)

            if not text_parts:
                continue

            name = " ".join(text_parts)

            # =========================================================
            # CLEAN DUPLICATED PATTERNS
            # =========================================================

            name = re.sub(r"\bX KG\b\s+\bX KG\b", "X KG", name, flags=re.I)

            name = re.sub(r"\b(\d+)\s*KG\b\s+\1\s*KG\b", r"\1 KG", name, flags=re.I)

            name = re.sub(r"\s+", " ", name).strip()

            normalized_name = normalize(name)

            # =========================================================
            # EXPLICIT BRAND
            # =========================================================

            explicit_brand = None

            for brand in sorted(CRITICAL_BRANDS, key=len, reverse=True):

                if normalize(brand) in normalized_name:
                    explicit_brand = brand.lower()
                    break

            # =========================================================
            # FALLBACK BRAND
            # =========================================================

            if current_brand in SECTION_HEADERS:
                fallback_brand = ""
            else:
                fallback_brand = current_brand

            final_brand = explicit_brand or fallback_brand

            # =========================================================
            # BUILD FULL NAME
            # =========================================================

            if final_brand:

                # avoid duplication:
                # "oreo OREO..."
                if normalize(final_brand) not in normalized_name:
                    full_name = f"{final_brand} {name}".strip()
                else:
                    full_name = name

            else:
                full_name = name

            products.append(
                {
                    "brand": final_brand,
                    "name": full_name,
                    "raw_name": name,
                    "price": price,
                }
            )

    return products
def parse_txt_local_products(path):

    products = []

    current_brand = ""

    with open(path, "r", encoding="utf-8") as f:

        for raw_line in f:

            raw = raw_line.strip()

            if not raw:
                continue

            if "$" not in raw:

                upper = raw.upper().replace(":", "").strip()

                if len(upper.split()) <= 4:
                    current_brand = upper
                    continue

            price_match = re.search(r"\$ ?([\d\.]+)", raw)

            if not price_match:
                continue

            try:
                price = int(price_match.group(1).replace(".", ""))
            except:
                continue

            name = raw[: price_match.start()].strip()

            normalized_name = normalize(name)

            explicit_brand = None

            for brand in CRITICAL_BRANDS:

                if normalize(brand) in normalized_name:
                    explicit_brand = brand
                    break

            final_brand = explicit_brand or current_brand.lower()

            full_name = f"{final_brand} {name}".strip()

            products.append(
                {
                    "brand": final_brand,
                    "name": full_name,
                    "raw_name": name,
                    "price": price,
                }
            )

    return products


# ============================================================
# FETCH WEB PRODUCTS
# ============================================================


def fetch_web_products():

    session = requests.Session()
    session.headers.update(HEADERS)

    all_products = []

    page = 1

    while True:

        payload = {
            "page": str(page),
            "sort_by": "created-at-descending",
            "per_page": "200",
        }

        r = session.post(
            f"{BASE_URL}/v3/products/advanced-search",
            json=payload,
        )

        print(f"{BLUE}[+] PAGE {page}: {r.status_code}{RESET}")

        if r.status_code != 200:
            print(r.text)
            break

        data = r.json()

        products = data.get("results", [])

        if not products:
            break

        for p in products:

            try:

                name = p.get("name", {}).get("es", "")

                handle = p.get("handle", {}).get("es", "")

                desc = p.get("description", {}).get("es", "")

                brand = p.get("brand") or ""

                variants = p.get("variants", [])

                if not variants:
                    continue

                valid_variants = [v for v in variants if v.get("price")]

                if not valid_variants:
                    continue

                variant = min(
                    valid_variants, key=lambda v: float(v.get("price", 999999))
                )

                full_text = " ".join(
                    [
                        name,
                        handle,
                        desc,
                        brand,
                    ]
                )

                all_products.append(
                    {
                        "id": p["id"],
                        "variant_id": variant["id"],
                        "name": name,
                        "full_text": full_text,
                        "price": int(float(variant["price"])),
                    }
                )

            except Exception as e:
                print("parse error:", e)

        page += 1

    return all_products


# ============================================================
# BLOCKED MATCH
# ============================================================


def blocked_match(reason):

    return {
        "score": -999,
        "status": "BLOCKED",
        "confidence": 0,
        "reasons": [reason],
    }


# ============================================================
# MATCH
# ============================================================


def compute_match(local, web):

    reasons = []

    local_name = normalize(local["name"])
    web_name = normalize(web["full_text"])

    local_tokens = set(tokenize(local_name))
    web_tokens = set(tokenize(web_name))

    local_cat = detect_category(local_name)
    web_cat = detect_category(web_name)

    if local_cat and web_cat and (local_cat, web_cat) in INVALID_CATEGORY_COMBINATIONS:
        return blocked_match(f"invalid category combination {local_cat}->{web_cat}")

    score = 0

    # ========================================================
    # TOKEN SCORE
    # ========================================================

    intersection = local_tokens & web_tokens

    token_score = 0

    for token in intersection:

        val = IMPORTANT_TOKENS.get(token, 10)

        token_score += val

        reasons.append(f"+{val} token:{token}")

    score += token_score

    # ========================================================
    # GENERIC WORD PROTECTION
    # ========================================================

    if len(intersection - GENERIC_WORDS) == 0:
        score -= 100
        reasons.append("-100 generic words only")

    # ========================================================
    # STRING SIMILARITY
    # ========================================================

    sim = similarity(local_name, web_name)

    sim_score = sim * 50

    score += sim_score

    reasons.append(f"+{sim_score:.2f} similarity")

    # ========================================================
    # BRAND
    # ========================================================

    local_brand = extract_brand(local_name)
    web_brand = extract_brand(web_name)

    brand_conflict = False

    if local_brand and web_brand:

        if local_brand == web_brand:

            score += 40
            reasons.append("+40 same brand")

        else:

            score -= 250
            reasons.append("-250 brand mismatch")

            brand_conflict = True

    # ========================================================
    # CATEGORY
    # ========================================================

    category_conflict = False

    if local_cat and web_cat:

        if local_cat == web_cat:

            score += 30
            reasons.append("+30 same category")

        else:

            score -= 150
            reasons.append("-150 category mismatch")

            category_conflict = True

    # ========================================================
    # FIAMBRE PROTECTION
    # ========================================================

    local_fiambre = FIAMBRE_TYPES & local_tokens
    web_fiambre = FIAMBRE_TYPES & web_tokens

    if local_fiambre and web_fiambre:

        if local_fiambre != web_fiambre:

            score -= 500

            reasons.append("-500 fiambre mismatch")

    # ========================================================
    # STRUCTURED UNITS
    # ========================================================

    local_units = extract_structured_units(local_name)
    web_units = extract_structured_units(web_name)

    weight_conflict = False

    if local_units["kg"] and web_units["kg"]:

        ratio = max(local_units["kg"], web_units["kg"]) / min(
            local_units["kg"], web_units["kg"]
        )

        if ratio > 3:

            score -= 120

            reasons.append("-120 kg mismatch")

            weight_conflict = True

        else:

            score += 20

            reasons.append("+20 kg compatible")

    if local_units["units"] and web_units["units"]:

        ratio = max(local_units["units"], web_units["units"]) / min(
            local_units["units"], web_units["units"]
        )

        if ratio > 3:

            score -= 100

            reasons.append("-100 unit mismatch")

            weight_conflict = True

    # ========================================================
    # PRICE
    # ========================================================

    local_target = local["price"] * 1.07

    price_diff = percent_diff(
        local_target,
        web["price"],
    )

    if price_diff > 0.60:

        score -= 120

        reasons.append("-120 huge price diff")

    elif price_diff > 0.30:

        score -= 60

        reasons.append("-60 large price diff")

    else:

        score += 15

        reasons.append("+15 compatible price")

    # ========================================================
    # CONFIDENCE
    # ========================================================

    confidence = min(max(score / 150, 0), 1)

    if confidence > 0.90:
        confidence_label = "VERY_HIGH"

    elif confidence > 0.75:
        confidence_label = "HIGH"

    elif confidence > 0.50:
        confidence_label = "MEDIUM"

    else:
        confidence_label = "LOW"

    # ========================================================
    # STATUS
    # ========================================================

    if score < 0:
        status = "BLOCKED"

    elif price_diff > 0.45 or weight_conflict or brand_conflict or category_conflict:
        status = "REVIEW"

    else:
        status = "SAFE"

    return {
        "score": score,
        "status": status,
        "confidence": confidence,
        "confidence_label": confidence_label,
        "price_diff": price_diff,
        "brand_conflict": brand_conflict,
        "category_conflict": category_conflict,
        "weight_conflict": weight_conflict,
        "reasons": reasons,
    }


# ============================================================
# BUILD CATEGORY INDEX
# ============================================================


def build_category_index(web_products):

    idx = defaultdict(list)

    for w in web_products:

        cat = detect_category(w["full_text"])

        idx[cat].append(w)

    return idx


# ============================================================
# BUILD MATCHES
# ============================================================


def build_matches(local_products, web_products):

    candidates = []

    category_index = build_category_index(web_products)

    for li, local in enumerate(local_products):

        local_cat = detect_category(local["name"])

        candidate_webs = category_index.get(local_cat, [])

        if local_cat is None or not candidate_webs:
            candidate_webs = web_products

        for wi, web in enumerate(candidate_webs):

            result = compute_match(local, web)

            if result["score"] < MIN_SCORE:
                continue

            candidates.append(
                {
                    "local_index": li,
                    "web": web,
                    "result": result,
                }
            )

    candidates.sort(
        key=lambda x: x["result"]["score"],
        reverse=True,
    )

    used_local = set()
    used_web = set()

    matches = []

    for c in candidates:

        li = c["local_index"]

        web_id = c["web"]["id"]

        if li in used_local:
            continue

        if web_id in used_web:
            continue

        used_local.add(li)
        used_web.add(web_id)

        matches.append(
            {
                "local": local_products[li],
                "web": c["web"],
                "result": c["result"],
            }
        )

    not_found = []

    for i, local in enumerate(local_products):

        if i not in used_local:
            not_found.append(local)

    return matches, not_found


# ============================================================
# PATCH PRICE
# ============================================================


def patch_price(session, product_id, variant_id, new_price):

    payload = {
        "variants": [
            {
                "id": int(variant_id),
                "price": float(new_price),
            }
        ]
    }

    r = session.patch(
        f"{BASE_URL}/v4/products/{product_id}",
        json=payload,
    )

    return r.status_code, r.text


# ============================================================
# MAIN
# ============================================================


def main():

    print(f"{BLUE}Loading local products...{RESET}")

    if LOCAL_FILE.endswith("txt"):
        local_products = parse_txt_local_products(LOCAL_FILE)
    elif LOCAL_FILE.endswith("xlsx"):
        local_products = parse_xlsx_local_products(LOCAL_FILE)
    else:
        print("File not suported")
        exit()

    print(f"{GREEN}Local: {len(local_products)}{RESET}")

    print()

    print(f"{BLUE}Fetching web products...{RESET}")

    web_products = fetch_web_products()

    print(f"{GREEN}Web: {len(web_products)}{RESET}")

    with open("web_products.json", "w", encoding="utf-8") as f:

        json.dump(
            web_products,
            f,
            ensure_ascii=False,
            indent=2,
        )

    print()

    print(f"{BLUE}Matching...{RESET}")

    matches, not_found = build_matches(
        local_products,
        web_products,
    )

    print()

    print(f"{MAGENTA}=============================={RESET}")
    print(f"{MAGENTA}NOT FOUND{RESET}")
    print(f"{MAGENTA}=============================={RESET}")

    for p in not_found:

        print(f"{RED}NOT FOUND:{RESET} " f"{p['name']} " f"${p['price']}")

    print()

    print(f"{MAGENTA}=============================={RESET}")
    print(f"{MAGENTA}MATCHES{RESET}")
    print(f"{MAGENTA}=============================={RESET}")

    session = requests.Session()
    session.headers.update(HEADERS)

    for i, m in enumerate(matches):

        local = m["local"]
        web = m["web"]
        result = m["result"]

        print()

        print(f"{CYAN}MATCH [{i}]{RESET}")

        print(f"{GREEN}LOCAL:{RESET} " f"{local['name']} " f"${local['price']}")

        print(f"{GREEN}WEB:{RESET} " f"{web['name']} " f"${web['price']}")

        print(
            f"score={result['score']:.2f} "
            f"status={result['status']} "
            f"confidence={result['confidence_label']}"
        )

        print()

        for reason in result["reasons"]:
            print("  ", reason)

        if int(local["price"] * 1.07) == web["price"]:

            print(f"{GREEN}Same price{RESET}")

            continue

        if result["status"] == "BLOCKED":

            print(f"{RED}BLOCKED{RESET}")

            continue

        if result["status"] == "SAFE" and AUTO_CONFIRM_SAFE:
            confirm = "y"

        else:

            confirm = input(f"{YELLOW}Change price? (y/n): {RESET}").strip().lower()

        if confirm != "y":
            continue

        print(f"{BLUE}Updating:{RESET} " f"${web['price']} -> ${local['price']}")

        # ====================================================
        # ENABLE WHEN READY
        # ====================================================

        status, text = patch_price(
            session,
            web["id"],
            web["variant_id"],
            int(local["price"] * 1.07),
        )

        print(f"PATCH STATUS: {status}")

        if status != 200:
            print(text)

    print()

    print(f"{GREEN}Done.{RESET}")


# ============================================================

if __name__ == "__main__":
    main()
