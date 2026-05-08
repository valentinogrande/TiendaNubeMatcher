#!/usr/bin/env python3

import base64
import difflib
import html
import io
import json
import os
import re
import shutil
import signal
import subprocess
import sys
import tempfile
import unicodedata
import uuid
from collections import defaultdict
from functools import lru_cache

import requests
from dotenv import load_dotenv
from openpyxl import load_workbook
from PIL import Image

# ============================================================
# CONFIG
# ============================================================

load_dotenv()


def signal_handler(sig, frame):

    print()
    print(f"{YELLOW}Interrupted by user (CTRL+C){RESET}")
    exit()


signal.signal(signal.SIGINT, signal_handler)

ACCESS_TOKEN = os.getenv("ACCESS_TOKEN")

STORE_ID = os.getenv("STORE_ID")
USER_ID = os.getenv("USER_ID")

BASE_URL = os.getenv("BASE_URL")

LOCAL_FILE = os.getenv("LOCAL_FILE")

AUTO_CONFIRM_SAFE = False

MIN_SCORE = 40

# ============================================================
# HEADERS
# ============================================================

HEADERS = {
    "x-admin-platform": "web",
    "x-admin-front-version": "2026.4.30",
    "x-admin-build-version": "2f2aead",
    "x-access-token": ACCESS_TOKEN,
    "x-store-id": STORE_ID,
    "x-user-id": USER_ID,
    "x-authentication-factor-code": "no-defined",
    "x-store-country": "AR",
    "Content-Type": "application/json",
    "User-Agent": (
        "Mozilla/5.0 (X11; Linux x86_64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/146.0.0.0 Safari/537.36"
    ),
}

# ============================================================
# COLORS
# ============================================================

RED = "\033[91m"
GREEN = "\033[92m"
YELLOW = "\033[93m"
BLUE = "\033[94m"
MAGENTA = "\033[95m"
CYAN = "\033[96m"
RESET = "\033[0m"

# ============================================================
# STOPWORDS
# ============================================================

STOPWORDS = {
    "x",
    "kg",
    "gr",
    "g",
    "de",
    "la",
    "el",
    "los",
    "las",
    "y",
    "con",
    "por",
    "unidad",
    "unidades",
    "pack",
    "u",
    "uni",
}

# ============================================================
# BRANDS
# ============================================================


CRITICAL_BRANDS = {
    "de la huerta",
    "paladini",
    "fela",
    "sadia",
    "la martina",
    "vasconia",
    "cremac",
    "pacama",
    "albor",
    "san cayetano",
    "rosamonte",
    "andresito",
    "playadito",
    "cbse",
    "terrabusi",
    "lucchetti",
    "galletti",
    "mocovi",
    "cabaña espinillo",
    "punta del agua",
    "lealtad",
    "molto",
    "arcor",
    "baggio",
    "mora",
    "milkey",
    "oreo",
    "aguila",
    "erevan",
    "veronica",
    "bon o bon",
    "pronatte",
    "regio",
    "elegante",
}

SECTION_HEADERS = {
    "cucharitas",
    "frutas secas",
    "leche industrial",
    "termicos",
    "mantecol",
    "chocolate",
    "chocolates aguila",
}


# ============================================================
# SYNONYMS
# ============================================================

SYNONYMS = {
    "muzza": "mozzarella",
    "mozza": "mozzarella",
    "jyq": "jamon queso",
    "paleta cocida": "paleta",
    "regianito": "reggiano",
}

# ============================================================
# IMPORTANT TOKENS
# ============================================================

IMPORTANT_TOKENS = {
    "reggiano": 40,
    "sardo": 35,
    "mortadela": 35,
    "jamon": 35,
    "paleta": 35,
    "salchicha": 35,
    "mozzarella": 40,
    "cheddar": 40,
    "romano": 40,
    "fontina": 40,
    "holanda": 40,
    "cremoso": 30,
    "salut": 30,
}

# ============================================================
# GENERIC WORDS
# ============================================================

GENERIC_WORDS = {
    "barra",
    "cremoso",
    "light",
}

# ============================================================
# FIAMBRE TYPES
# ============================================================

FIAMBRE_TYPES = {
    "jamon",
    "paleta",
    "mortadela",
    "salchicha",
    "milan",
}

# ============================================================
# CATEGORY KEYWORDS
# ============================================================

CATEGORY_KEYWORDS = {
    "queso": [
        "cremoso",
        "mozzarella",
        "barra",
        "sardo",
        "fontina",
        "holanda",
        "danbo",
        "reggiano",
        "romano",
        "por salut",
        "cheddar",
    ],
    "limpieza": [
        "detergente",
        "lavandina",
        "jabon",
    ],
    "yerba": [
        "yerba",
        "mate",
        "rosamonte",
        "playadito",
        "cbse",
        "andresito",
    ],
    "fideos": [
        "fideos",
        "spaghetti",
        "tirabuzon",
        "mostachol",
    ],
    "arroz": [
        "arroz",
    ],
    "tomate": [
        "tomate",
        "pure",
        "salsa",
    ],
    "aceitunas": [
        "aceituna",
        "aceitunas",
    ],
    "fiambre": [
        "mortadela",
        "jamon",
        "paleta",
        "salchicha",
        "milan",
    ],
    "lacteo": [
        "leche",
        "crema",
    ],
}

# ============================================================
# INVALID CATEGORY COMBINATIONS
# ============================================================

INVALID_CATEGORY_COMBINATIONS = {
    ("queso", "limpieza"),
    ("yerba", "fiambre"),
    ("aceitunas", "fideos"),
    ("tomate", "limpieza"),
}

# ============================================================
# NORMALIZE
# ============================================================


@lru_cache(maxsize=50000)
def normalize(text):

    text = unicodedata.normalize("NFKD", text)
    text = text.encode("ascii", "ignore").decode("ascii")

    text = text.lower()

    for k, v in SYNONYMS.items():
        text = text.replace(k, v)

    text = re.sub(r"[^\w\s]", " ", text)

    text = re.sub(r"\s+", " ", text)

    return text.strip()


# ============================================================
# TOKENIZE
# ============================================================


def tokenize(text):

    text = normalize(text)

    tokens = []

    for t in text.split():

        if t in STOPWORDS:
            continue

        t = re.sub(r"^(\d+)(?:kg|g|u|uni|unid|lt|l)$", r"\1", t, flags=re.I)

        if len(t) > 4 and t.endswith("es"):
            t = t[:-2]
        elif len(t) > 3 and t.endswith("s"):
            t = t[:-1]

        tokens.append(t)

    return tokens


# ============================================================
# SIMILARITY
# ============================================================


def similarity(a, b):

    return difflib.SequenceMatcher(None, a, b).ratio()


# ============================================================
# CATEGORY
# ============================================================


def detect_category(text):

    t = normalize(text)

    for category, kws in CATEGORY_KEYWORDS.items():

        for kw in kws:

            if kw in t:
                return category

    return None


# ============================================================
# BRAND
# ============================================================


def extract_brand(text):

    t = normalize(text)

    found = []

    for brand in CRITICAL_BRANDS:

        if normalize(brand) in t:
            found.append(brand)

    if not found:
        return None

    found.sort(key=len, reverse=True)

    return found[0]


# ============================================================
# STRUCTURED UNITS
# ============================================================


def extract_structured_units(text):

    t = normalize(text)

    result = {
        "kg": None,
        "g": None,
        "units": None,
    }

    kg = re.findall(r"(\d+(?:\.\d+)?)\s*kg", t)

    if kg:
        result["kg"] = float(kg[0])

    g = re.findall(r"(\d+(?:\.\d+)?)\s*g", t)

    if g:
        result["g"] = float(g[0])

    units = re.findall(r"\bx\s*(\d+(?:\.\d+)?)", t)

    if units:
        result["units"] = int(units[0])

    return result


# ============================================================
# PRICE DIFF
# ============================================================


def percent_diff(a, b):

    if a <= 0 or b <= 0:
        return 999

    return abs(a - b) / max(a, b)


# ============================================================
# PARSE LOCAL PRODUCTS
# ============================================================


def parse_xlsx_local_products(path):

    products = []

    wb = load_workbook(path, data_only=True)

    for sheet_name in wb.sheetnames:
        if sheet_name == "Hoja1":
            continue
        ws = wb[sheet_name]

        current_brand = ""

        for row in ws.iter_rows(values_only=True):

            values = [v for v in row if v is not None]

            if not values:
                continue

            # =========================================================
            # HEADER DETECTION
            # =========================================================

            first = str(values[0]).strip()

            if (
                isinstance(values[0], str)
                and len(values) == 1
                and len(first.split()) <= 4
            ):

                current_brand = first.replace('"', "").replace(":", "").strip().lower()

                continue

            # =========================================================
            # PRICE
            # =========================================================

            price = None

            for v in reversed(values):

                parsed = None

                # =====================================================
                # NUMBERS
                # =====================================================

                if isinstance(v, (int, float)):

                    if v < 100:
                        continue

                    parsed = int(round(v))

                # =====================================================
                # STRINGS
                # =====================================================

                elif isinstance(v, str):

                    s = v.strip()

                    if not re.search(r"\d", s):
                        continue

                    # remove currency
                    s = s.replace("$", "")

                    # remove spaces
                    s = s.replace(" ", "")

                    # 49,300.00
                    if "," in s and "." in s:

                        if s.rfind(",") < s.rfind("."):
                            s = s.replace(",", "")

                    # 49.300,00
                    elif "," in s:

                        s = s.replace(".", "")
                        s = s.replace(",", ".")

                    try:
                        parsed = int(round(float(s)))
                    except:
                        continue

                # =====================================================
                # VALIDATE
                # =====================================================

                if parsed and parsed >= 100:
                    price = parsed
                    break
            # =========================================================
            # TEXT PARTS
            # =========================================================

            seen = set()

            text_parts = []

            for v in values:

                if isinstance(v, (int, float)):
                    continue

                s = str(v).strip()

                if not s:
                    continue

                norm = normalize(s)

                # remove duplicate columns
                if norm in seen:
                    continue

                seen.add(norm)

                text_parts.append(s)

            if not text_parts:
                continue

            name = " ".join(text_parts)

            # =========================================================
            # CLEAN DUPLICATED PATTERNS
            # =========================================================

            name = re.sub(r"\bX KG\b\s+\bX KG\b", "X KG", name, flags=re.I)

            name = re.sub(r"\b(\d+)\s*KG\b\s+\1\s*KG\b", r"\1 KG", name, flags=re.I)

            name = re.sub(r"\s+", " ", name).strip()

            normalized_name = normalize(name)

            # =========================================================
            # EXPLICIT BRAND
            # =========================================================

            explicit_brand = None

            for brand in sorted(CRITICAL_BRANDS, key=len, reverse=True):

                if normalize(brand) in normalized_name:
                    explicit_brand = brand.lower()
                    break

            # =========================================================
            # FALLBACK BRAND
            # =========================================================

            if current_brand in SECTION_HEADERS:
                fallback_brand = ""
            else:
                fallback_brand = current_brand

            final_brand = explicit_brand or fallback_brand

            # =========================================================
            # BUILD FULL NAME
            # =========================================================

            if final_brand:

                # avoid duplication:
                # "oreo OREO..."
                if normalize(final_brand) not in normalized_name:
                    full_name = f"{final_brand} {name}".strip()
                else:
                    full_name = name

            else:
                full_name = name

            if price is None:
                continue

            products.append(
                {
                    "brand": final_brand,
                    "name": full_name,
                    "raw_name": name,
                    "price": price,
                }
            )

    return products


def parse_txt_local_products(path):

    products = []

    current_brand = ""

    with open(path, "r", encoding="utf-8") as f:

        for raw_line in f:

            raw = raw_line.strip()

            if not raw:
                continue

            if "$" not in raw:

                upper = raw.upper().replace(":", "").strip()

                if len(upper.split()) <= 4:
                    current_brand = upper
                    continue

            price_match = re.search(r"\$ ?([\d\.]+)", raw)

            if not price_match:
                continue

            try:
                price = int(price_match.group(1).replace(".", ""))
            except:
                continue

            name = raw[: price_match.start()].strip()

            normalized_name = normalize(name)

            explicit_brand = None

            for brand in CRITICAL_BRANDS:

                if normalize(brand) in normalized_name:
                    explicit_brand = brand
                    break

            final_brand = explicit_brand or current_brand.lower()

            full_name = f"{final_brand} {name}".strip()

            products.append(
                {
                    "brand": final_brand,
                    "name": full_name,
                    "raw_name": name,
                    "price": price,
                }
            )

    return products


# ============================================================
# FETCH WEB PRODUCTS
# ============================================================


def fetch_web_products():

    session = requests.Session()
    session.headers.update(HEADERS)

    all_products = []

    page = 1

    while True:

        payload = {
            "page": str(page),
            "sort_by": "created-at-descending",
            "per_page": "200",
        }

        r = session.post(
            f"{BASE_URL}/v3/products/advanced-search",
            json=payload,
        )

        print(f"{BLUE}[+] PAGE {page}: {r.status_code}{RESET}")

        if r.status_code != 200:
            print(r.text)
            break

        data = r.json()

        products = data.get("results", [])

        if not products:
            break

        for p in products:

            try:

                name = p.get("name", {}).get("es", "")

                handle = p.get("handle", {}).get("es", "")

                desc = p.get("description", {}).get("es", "")

                brand = p.get("brand") or ""

                variants = p.get("variants", [])

                if not variants:
                    continue

                valid_variants = [v for v in variants if v.get("price")]

                if not valid_variants:
                    continue

                variant = min(
                    valid_variants, key=lambda v: float(v.get("price", 999999))
                )

                full_text = " ".join(
                    [
                        name,
                        handle,
                        desc,
                        brand,
                    ]
                )

                all_products.append(
                    {
                        "id": p["id"],
                        "variant_id": variant["id"],
                        "name": name,
                        "full_text": full_text,
                        "price": int(float(variant["price"])),
                    }
                )

            except Exception as e:
                print("parse error:", e)

        page += 1

    return all_products


# ============================================================
# BLOCKED MATCH
# ============================================================


def blocked_match(reason):

    return {
        "score": -999,
        "status": "BLOCKED",
        "confidence": 0,
        "reasons": [reason],
    }


# ============================================================
# MATCH
# ============================================================


def compute_match(local, web):

    reasons = []

    local_name = normalize(local["name"])
    web_name = normalize(web["full_text"])

    local_tokens = set(tokenize(local_name))
    web_tokens = set(tokenize(web_name))

    local_cat = detect_category(local_name)
    web_cat = detect_category(web_name)

    if local_cat and web_cat and (local_cat, web_cat) in INVALID_CATEGORY_COMBINATIONS:
        return blocked_match(f"invalid category combination {local_cat}->{web_cat}")

    score = 0

    # ========================================================
    # TOKEN SCORE
    # ========================================================

    intersection = local_tokens & web_tokens

    token_score = 0

    for token in intersection:

        val = IMPORTANT_TOKENS.get(token, 10)

        token_score += val

        reasons.append(f"+{val} token:{token}")

    score += token_score

    # ========================================================
    # GENERIC WORD PROTECTION
    # ========================================================

    if len(intersection - GENERIC_WORDS) == 0:
        score -= 100
        reasons.append("-100 generic words only")

    # ========================================================
    # STRING SIMILARITY
    # ========================================================

    sim = similarity(local_name, web_name)

    sim_score = sim * 50

    score += sim_score

    reasons.append(f"+{sim_score:.2f} similarity")

    # ========================================================
    # BRAND
    # ========================================================

    local_brand = extract_brand(local_name)
    web_brand = extract_brand(web_name)

    brand_conflict = False

    if local_brand and web_brand:

        if local_brand == web_brand:

            score += 40
            reasons.append("+40 same brand")

        else:

            score -= 250
            reasons.append("-250 brand mismatch")

            brand_conflict = True

    # ========================================================
    # CATEGORY
    # ========================================================

    category_conflict = False

    if local_cat and web_cat:

        if local_cat == web_cat:

            score += 30
            reasons.append("+30 same category")

        else:

            score -= 150
            reasons.append("-150 category mismatch")

            category_conflict = True

    # ========================================================
    # FIAMBRE PROTECTION
    # ========================================================

    local_fiambre = FIAMBRE_TYPES & local_tokens
    web_fiambre = FIAMBRE_TYPES & web_tokens

    if local_fiambre and web_fiambre:

        if local_fiambre != web_fiambre:

            score -= 500

            reasons.append("-500 fiambre mismatch")

    # ========================================================
    # STRUCTURED UNITS
    # ========================================================

    local_units = extract_structured_units(local_name)
    web_units = extract_structured_units(web_name)

    weight_conflict = False

    if local_units["kg"] and web_units["kg"]:

        ratio = max(local_units["kg"], web_units["kg"]) / min(
            local_units["kg"], web_units["kg"]
        )

        if ratio > 3:

            score -= 120

            reasons.append("-120 kg mismatch")

            weight_conflict = True

        else:

            score += 20

            reasons.append("+20 kg compatible")

    if local_units["units"] and web_units["units"]:

        ratio = max(local_units["units"], web_units["units"]) / min(
            local_units["units"], web_units["units"]
        )

        if ratio > 3:

            score -= 100

            reasons.append("-100 unit mismatch")

            weight_conflict = True

    # ========================================================
    # PRICE
    # ========================================================

    local_target = local["price"] * 1.07

    price_diff = percent_diff(
        local_target,
        web["price"],
    )

    if price_diff > 0.60:

        score -= 120

        reasons.append("-120 huge price diff")

    elif price_diff > 0.30:

        score -= 60

        reasons.append("-60 large price diff")

    else:

        score += 15

        reasons.append("+15 compatible price")

    # ========================================================
    # CONFIDENCE
    # ========================================================

    confidence = min(max(score / 150, 0), 1)

    if confidence > 0.90:
        confidence_label = "VERY_HIGH"

    elif confidence > 0.75:
        confidence_label = "HIGH"

    elif confidence > 0.50:
        confidence_label = "MEDIUM"

    else:
        confidence_label = "LOW"

    # ========================================================
    # STATUS
    # ========================================================

    if score < 0:
        status = "BLOCKED"

    elif price_diff > 0.45 or weight_conflict or brand_conflict or category_conflict:
        status = "REVIEW"

    else:
        status = "SAFE"

    return {
        "score": score,
        "status": status,
        "confidence": confidence,
        "confidence_label": confidence_label,
        "price_diff": price_diff,
        "brand_conflict": brand_conflict,
        "category_conflict": category_conflict,
        "weight_conflict": weight_conflict,
        "reasons": reasons,
    }


# ============================================================
# BUILD CATEGORY INDEX
# ============================================================


def build_category_index(web_products):

    idx = defaultdict(list)

    for w in web_products:

        cat = detect_category(w["full_text"])

        idx[cat].append(w)

    return idx


# ============================================================
# BUILD MATCHES
# ============================================================


def build_matches(local_products, web_products):

    candidates = []

    category_index = build_category_index(web_products)

    for li, local in enumerate(local_products):

        local_cat = detect_category(local["name"])

        candidate_webs = category_index.get(local_cat, [])

        if local_cat is None or not candidate_webs:
            candidate_webs = web_products

        for wi, web in enumerate(candidate_webs):

            result = compute_match(local, web)

            if result["score"] < MIN_SCORE:
                continue

            candidates.append(
                {
                    "local_index": li,
                    "web": web,
                    "result": result,
                }
            )

    candidates.sort(
        key=lambda x: x["result"]["score"],
        reverse=True,
    )

    used_local = set()
    used_web = set()

    matches = []

    for c in candidates:

        li = c["local_index"]

        web_id = c["web"]["id"]

        if li in used_local:
            continue

        if web_id in used_web:
            continue

        used_local.add(li)
        used_web.add(web_id)

        matches.append(
            {
                "local": local_products[li],
                "web": c["web"],
                "result": c["result"],
            }
        )

    not_found = []

    for i, local in enumerate(local_products):

        if i not in used_local:
            not_found.append(local)

    for prod in not_found:

        name_lower = prod["name"].lower()

        prod["brand"] = ""

        for brand in CRITICAL_BRANDS:

            if brand.lower() in name_lower:
                prod["brand"] = brand
                break

    return matches, not_found


# ============================================================
# PATCH PRICE
# ============================================================


def patch_price(session, product_id, variant_id, new_price):

    payload = {
        "variants": [
            {
                "id": int(variant_id),
                "price": float(new_price),
            }
        ]
    }

    r = session.patch(
        f"{BASE_URL}/v4/products/{product_id}",
        json=payload,
    )

    return r.status_code, r.text


# ============================================================
# TERMINAL IMAGE DISPLAY
# ============================================================

TERM = os.environ.get("TERM", "")
TERM_PROGRAM = os.environ.get("TERM_PROGRAM", "")


def _supports_kitty() -> bool:
    return "kitty" in TERM


def _supports_iterm() -> bool:
    return TERM_PROGRAM in ("iTerm.app", "WezTerm") or "iterm" in TERM.lower()


def _supports_sixel() -> bool:
    # chafa con sixel como último recurso si está instalado
    return shutil.which("chafa") is not None


def _print_kitty(data: bytes):
    b64 = base64.standard_b64encode(data).decode()
    # Kitty APC: chunk de hasta 4096 bytes
    chunk = 4096
    chunks = [b64[i : i + chunk] for i in range(0, len(b64), chunk)]
    for idx, c in enumerate(chunks):
        m = 1 if idx < len(chunks) - 1 else 0
        first = idx == 0
        if first:
            header = f"\x1b_Ga=T,f=100,m={m};"
        else:
            header = f"\x1b_Gm={m};"
        sys.stdout.buffer.write(f"{header}{c}\x1b\\".encode())
    sys.stdout.buffer.write(b"\n")
    sys.stdout.buffer.flush()


def _print_iterm(data: bytes):
    b64 = base64.standard_b64encode(data).decode()
    length = len(data)
    sys.stdout.buffer.write(
        f"\x1b]1337;File=inline=1;size={length};width=40;height=20:{b64}\a\n".encode()
    )
    sys.stdout.buffer.flush()


def _print_chafa(data: bytes):
    with tempfile.NamedTemporaryFile(suffix=".jpg", delete=False) as f:
        f.write(data)
        tmp = f.name
    try:
        subprocess.run(
            ["chafa", "--size=40x20", tmp],
            check=False,
        )
    finally:
        os.unlink(tmp)


def print_image_in_terminal(image_url: str) -> bool:
    """
    Descarga y renderiza la imagen en terminal.
    Retorna True si logró mostrarla.
    """
    try:
        r = requests.get(image_url, timeout=8)
        r.raise_for_status()
        data = r.content
    except Exception as e:
        print(f"{RED}  [IMG] No se pudo descargar: {e}{RESET}")
        return False

    if _supports_kitty():
        _print_kitty(data)
        return True
    elif _supports_iterm():
        _print_iterm(data)
        return True
    elif _supports_sixel():
        _print_chafa(data)
        return True
    else:
        print(f"{YELLOW}  [IMG] Terminal sin soporte gráfico. URL: {image_url}{RESET}")
        return False


# ============================================================
# GOOGLE IMAGE SEARCH FALLBACK
# ============================================================
def open_google_images(query: str):
    encoded = requests.utils.quote(query)
    url = f"https://www.google.com/search?tbm=isch&q={encoded}"
    import webbrowser

    webbrowser.open(url)
    print(f"{CYAN}  [IMG] Abriendo Google Images: {query!r}{RESET}")


# ============================================================
# IMAGE SEARCH  (OpenFoodFacts)
# ============================================================
def search_image(name: str, brand: str, barcode: str = "") -> str:
    session = requests.Session()
    session.headers.update({"User-Agent": "MCDistApp/1.0"})

    clean_barcode = re.sub(r"\D", "", str(barcode))
    if len(clean_barcode) >= 8:
        url = f"https://world.openfoodfacts.org/api/v0/product/{clean_barcode}.json"
        try:
            r = session.get(url, timeout=8)
            data = r.json()
            if data.get("status") == 1:
                prod = data.get("product", {})
                img = prod.get("image_front_url") or prod.get("image_url", "")
                if img:
                    return img
        except Exception:
            pass

    query = f"{name} {brand}".strip() if brand else name
    encoded = requests.utils.quote(query)
    search_url = (
        "https://world.openfoodfacts.org/cgi/search.pl"
        f"?search_terms={encoded}&search_simple=1&json=1&page_size=1"
    )
    try:
        r = session.get(search_url, timeout=8)
        data = r.json()
        products = data.get("products", [])
        if products:
            p = products[0]
            img = p.get("image_front_url") or p.get("image_url", "")
            if img:
                return img
    except Exception:
        pass

    return ""


# ============================================================
# INSERT PROD
# ============================================================

INSERT_ENDPOINT = f"{BASE_URL}/v4/products?sync-refresh=true"
IMAGE_UPLOAD_ENDPOINT = f"{BASE_URL}/v1/products/images/binary"
LOCATION_ID = os.getenv("LOCATION_ID")  # 01JPMSFX3PD8JE5P3MK60P0Q5G


def upload_image(image_url: str) -> int | None:
    """
    Descarga la imagen y la sube a Tiendanube como binario.
    Retorna el id de imagen o None si falla.
    """

    try:
        r = requests.get(image_url, timeout=8)
        r.raise_for_status()
        img_bytes = r.content
        content_type = r.headers.get("Content-Type", "image/jpeg").split(";")[0]
        ext_map = {
            "image/jpeg": "jpg",
            "image/jpg": "jpg",
            "image/png": "png",
            "image/webp": "webp",
            "image/gif": "gif",
        }
        raw_filename = image_url.split("/")[-1].split("?")[0]
        if "." in raw_filename:
            filename = raw_filename
        else:
            ext = ext_map.get(content_type, "jpg")
            filename = f"product.{ext}"

        # Tiendanube no acepta webp → convertir a JPEG
        if content_type == "image/webp" or filename.lower().endswith(".webp"):
            print(f"{YELLOW}  [IMG] Converting webp to JPEG...{RESET}")
            img = Image.open(io.BytesIO(img_bytes)).convert("RGB")
            buf = io.BytesIO()
            img.save(buf, format="JPEG", quality=90)
            img_bytes = buf.getvalue()
            content_type = "image/jpeg"
            filename = filename.rsplit(".", 1)[0] + ".jpg"

        print(
            f"{YELLOW}  [IMG] Uploading: filename={filename!r} content_type={content_type!r} size={len(img_bytes)}{RESET}"
        )
    except Exception as e:
        print(f"{RED}  [IMG] No se pudo descargar: {e}{RESET}")
        return None

    upload_headers = {k: v for k, v in HEADERS.items() if k.lower() != "content-type"}
    session = requests.Session()
    session.headers.update(upload_headers)
    resp = session.post(
        IMAGE_UPLOAD_ENDPOINT,
        files={"file": (filename, img_bytes, content_type)},
    )
    if resp.status_code == 201:
        img_id = int(resp.json()["id"])
        print(f"{GREEN}  [IMG] Subida OK id={img_id}{RESET}")
        return img_id
    else:
        print(
            f"{RED}  [IMG] Error subiendo imagen {resp.status_code}: {resp.text}{RESET}"
        )
        return None


# ============================================================
# SEO ENGINE
# ============================================================

RE_KG = re.compile(r"(?i)(\d+(?:[\.,]\d+)?)\s*(kg|kilo|kilos)")
RE_G = re.compile(r"(?i)(\d+(?:[\.,]\d+)?)\s*(g|gr|grs|gramos)")
RE_L = re.compile(r"(?i)(\d+(?:[\.,]\d+)?)\s*(l|lt|lts|litro|litros)")
RE_ML = re.compile(r"(?i)(\d+(?:[\.,]\d+)?)\s*(ml)")
RE_CC = re.compile(r"(?i)(\d+(?:[\.,]\d+)?)\s*(cc)")
RE_UNIT = re.compile(r"(?i)x\s*(\d+)\s*u")

STOPWORDS = {
    "x",
    "de",
    "del",
    "la",
    "el",
    "en",
    "y",
    "con",
    "para",
    "por",
    "al",
    "las",
    "los",
}


def normalize_seo(text: str) -> str:
    return (
        text.lower()
        .replace("á", "a")
        .replace("é", "e")
        .replace("í", "i")
        .replace("ó", "o")
        .replace("ú", "u")
        .replace("ñ", "n")
        .replace("ü", "u")
    )


def clean_name(name: str) -> str:
    out = []

    for word in name.split():
        lower = word.lower()

        if lower in STOPWORDS:
            out.append(lower)
            continue

        units = {
            "kg": "Kg",
            "gr": "Gr",
            "g": "Gr",
            "lt": "Lt",
            "lts": "Lts",
            "ml": "Ml",
            "cc": "Cc",
        }

        if lower in units:
            out.append(units[lower])
            continue

        out.append(lower.capitalize())

    return " ".join(out)


def extract_presentation(name: str):
    n = name.lower()

    patterns = [
        (RE_KG, "Kg"),
        (RE_G, "gr"),
        (RE_ML, "ml"),
        (RE_CC, "cc"),
        (RE_L, "Lt"),
    ]

    for regex, unit in patterns:
        m = regex.search(n)
        if m:
            return f"{m.group(1)} {unit}"

    m = RE_UNIT.search(n)
    if m:
        return f"{m.group(1)} unidades"

    if "pack" in n:
        return "Pack"

    if "caja" in n:
        return "Caja"

    return None


def rubric_keywords(rubric: str) -> str:
    r = rubric.lower()

    if "aceite" in r:
        return "aceite mayorista,aceite por mayor"
    elif "gallet" in r:
        return "galletitas mayorista,galletitas por mayor"
    elif "chocolate" in r:
        return "chocolate mayorista,chocolate para reposteria"
    elif "bebida" in r:
        return "bebidas mayorista,bebidas por mayor"
    elif "snack" in r:
        return "snacks mayorista,snacks por mayor"
    else:
        return "alimentos mayoristas,distribuidor de alimentos"


def rubric_audience(rubric: str) -> str:
    r = rubric.lower()

    if "helado" in r:
        return "heladerías artesanales y reposterías"

    if "bebida" in r:
        return "kioscos, almacenes y gastronomía"

    if "snack" in r:
        return "kioscos y autoservicios"

    return "almacenes, y comercios"


def rubric_use_case(rubric: str) -> str:
    r = rubric.lower()

    if "aceite" in r:
        return "Ideal para fritura y gastronomía."

    if "snack" in r:
        return "Producto de alta rotación para kioscos."

    if "bebida" in r:
        return "Ideal para reventa en comercios."

    return "Producto ideal para reventa y gastronomía."


def generate_slug(name: str, brand: str) -> str:
    base = f"{name} {brand}".strip()

    slug = normalize_seo(base)

    slug = re.sub(r"[^a-z0-9\s-]", "", slug)
    slug = re.sub(r"\s+", "-", slug)
    slug = re.sub(r"-+", "-", slug)

    return slug.strip("-")


def build_title(name: str, brand: str) -> str:
    if brand:
        title = f"{name} {brand} Precio Mayorista | MC Distribuidora"
    else:
        title = f"{name} Precio Mayorista | MC Distribuidora"

    return title[:60]


def build_description(name, brand, presentation, audience):
    text = (
        f"Comprá {name} {brand} {presentation or ''} "
        f"al mejor precio mayorista. "
        f"Envíos a todo el país. "
        f"Ideal para {audience}."
    )

    return text[:155]


def build_tags(name, brand, rubric):
    tags = set()

    tags.add(name.lower())

    if brand:
        tags.add(brand.lower())
        tags.add(f"{name.lower()} {brand.lower()}")

    tags.add(f"{name.lower()} mayorista")
    tags.add(f"{name.lower()} precio")

    extra = rubric_keywords(rubric)

    for t in extra.split(","):
        tags.add(t.strip())

    final = []
    total = 0

    for tag in tags:
        if total + len(tag) > 220:
            break

        final.append(tag)
        total += len(tag)

    return final


def build_html_description(name, brand, presentation, rubric, audience, use_case):
    name = html.escape(name)
    brand = html.escape(brand)
    rubric = html.escape(rubric)
    audience = html.escape(audience)
    use_case = html.escape(use_case)

    presentation_html = ""

    if presentation:
        presentation_html = (
            f"<li><strong>Presentación:</strong> " f"{html.escape(presentation)}</li>"
        )

    brand_html = ""

    if brand:
        brand_html = f"<li><strong>Marca:</strong> {brand}</li>"

    return f"""
<h2>{name} {brand}</h2>

<p>
Conseguí <strong>{name}</strong> al mejor precio mayorista.
{use_case}
</p>

<h3>Características</h3>

<ul>
{brand_html}
{presentation_html}
<li><strong>Categoría:</strong> {rubric}</li>
<li><strong>Envíos:</strong> A todo el país</li>
<li><strong>Venta:</strong> Mayorista y minorista</li>
</ul>

<h3>Ideal para</h3>

<p>{audience}</p>

<p>
<strong>MC Distribuidora</strong>
</p>
"""


# ============================================================
# INSERT SEO
# ============================================================


def insert_seo(product_id, prod_name, brand="", rubric="Varios"):
    name_clean = clean_name(prod_name)

    presentation = extract_presentation(prod_name)

    audience = rubric_audience(rubric)

    use_case = rubric_use_case(rubric)

    seo_title = build_title(name_clean, brand)

    seo_description = build_description(name_clean, brand, presentation, audience)

    tags = build_tags(name_clean, brand, rubric)

    html_description = build_html_description(
        name_clean, brand, presentation, rubric, audience, use_case
    )

    slug = generate_slug(name_clean, brand)

    payload = {
        "seo_title": {
            "es": seo_title,
            "default": seo_title,
        },
        "seo_description": {
            "es": seo_description,
            "default": seo_description,
        },
        "brand": brand.lower(),
        "tags": [{"tag": t} for t in tags[:10]],
    }

    url = f"{BASE_URL}/v4/products/{product_id}?sync-refresh=true"

    session = requests.Session()

    session.headers.update(HEADERS)

    session.headers["x-idempotency-key"] = str(uuid.uuid4())

    r = session.patch(url, json=payload)

    if r.status_code in (200, 201):
        print(f"{GREEN}[SEO] SEO OK{RESET}")
    else:
        print(f"{RED}[SEO] ERROR {r.status_code}:{RESET}")
        print(r.text)


def insert_prod(
    prod_name: str, prod_price: int, brand: str = "", barcode: str = "", rubric="Varios"
):
    insert = input("Queres agregar producto? y/n ")
    if insert != "y":
        return
    print(f"{BLUE}  [IMG] Buscando imagen para: {prod_name!r}...{RESET}")
    image_url = search_image(prod_name, brand, barcode)

    if image_url:
        print(f"{GREEN}  [IMG] Encontrada:{RESET} {image_url}")
        print_image_in_terminal(image_url)
        answer = (
            input(f"{YELLOW}  ¿Aceptar imagen? (y/n/g=Google): {RESET}").strip().lower()
        )
        if answer == "g":
            open_google_images(f"{prod_name} {brand}".strip())
            image_url = input(
                f"{YELLOW}  Pegá la URL de la imagen (enter para ninguna): {RESET}"
            ).strip()
        elif answer != "y":
            image_url = ""
    else:
        print(f"{YELLOW}  [IMG] No encontrada en OpenFoodFacts.{RESET}")
        answer = (
            input(f"{YELLOW}  ¿Buscar en Google Images? (y/n): {RESET}").strip().lower()
        )
        if answer == "y":
            open_google_images(f"{prod_name} {brand}".strip())
            image_url = input(
                f"{YELLOW}  Pegá la URL de la imagen (enter para ninguna): {RESET}"
            ).strip()

    # --- subir imagen primero ---
    image_id = None
    if image_url:
        image_id = upload_image(image_url)

    if not image_id:
        return

    payload = {
        "name": {"es": prod_name, "default": prod_name},
        "description": {"default": ""},
        "publish": True,
        "variants": [
            {
                "price": float(int(prod_price) * 1.07),
                "attributes": [],
                "weight": 0,
                "width": 0,
                "height": 0,
                "depth": 0,
                "visible": True,
                "order": 1,
                "inventory_levels": [{"location_id": LOCATION_ID, "stock": None}],
                "cost": 0,
                "promotional_price": 0,
                "metafields": [],
            }
        ],
    }
    if image_id:
        payload["images"] = [{"id": image_id, "order": 0, "alt": {"default": ""}}]

    session = requests.Session()
    session.headers.update(HEADERS)
    session.headers["x-idempotency-key"] = str(uuid.uuid4())
    r = session.post(INSERT_ENDPOINT, json=payload)
    if not r.status_code in (200, 201):
        print(f"{RED}  [INSERT] Error {r.status_code}: {r.text}{RESET}")
        return None

    data = r.json()

    product_id = data["id"]

    print(f"{GREEN}[INSERT] OK id={product_id}{RESET}")

    # ========================================================
    # AUTO SEO INSERT
    # ========================================================

    insert_seo(
        product_id=product_id,
        prod_name=prod_name,
        brand=brand,
        rubric=rubric,
    )

    return product_id


# ===========================================================
# MAIN
# ============================================================


def main():

    print(f"{BLUE}Loading local products...{RESET}")

    if LOCAL_FILE.endswith("txt"):
        local_products = parse_txt_local_products(LOCAL_FILE)
    elif LOCAL_FILE.endswith("xlsx"):
        local_products = parse_xlsx_local_products(LOCAL_FILE)
    else:
        print("File not suported")
        exit()

    print(f"{GREEN}Local: {len(local_products)}{RESET}")

    print()

    print(f"{BLUE}Fetching web products...{RESET}")

    web_products = fetch_web_products()

    print(f"{GREEN}Web: {len(web_products)}{RESET}")

    with open("web_products.json", "w", encoding="utf-8") as f:

        json.dump(
            web_products,
            f,
            ensure_ascii=False,
            indent=2,
        )

    print()

    print(f"{BLUE}Matching...{RESET}")

    matches, not_found = build_matches(
        local_products,
        web_products,
    )

    print()

    print(f"{MAGENTA}=============================={RESET}")
    print(f"{MAGENTA}NOT FOUND{RESET}")
    print(f"{MAGENTA}=============================={RESET}")

    for p in not_found:

        print(
            f"{RED}NOT FOUND:{RESET} " f"{p['name']} " f"${p['price']}" f"${p['brand']}"
        )

        insert_prod(p["name"], p["price"], p["brand"])

    print()

    print(f"{MAGENTA}=============================={RESET}")
    print(f"{MAGENTA}MATCHES{RESET}")
    print(f"{MAGENTA}=============================={RESET}")

    session = requests.Session()
    session.headers.update(HEADERS)

    for i, m in enumerate(matches):

        local = m["local"]
        web = m["web"]
        result = m["result"]

        print()

        print(f"{CYAN}MATCH [{i}]{RESET}")

        print(f"{GREEN}LOCAL:{RESET} " f"{local['name']} " f"${local['price']}")

        print(f"{GREEN}WEB:{RESET} " f"{web['name']} " f"${web['price']}")

        print(
            f"score={result['score']:.2f} "
            f"status={result['status']} "
            f"confidence={result['confidence_label']}"
        )

        print()

        for reason in result["reasons"]:
            print("  ", reason)

        if int(local["price"] * 1.07) == web["price"]:

            print(f"{GREEN}Same price{RESET}")

            continue

        if result["status"] == "BLOCKED":

            print(f"{RED}BLOCKED{RESET}")

            continue

        if result["status"] == "SAFE" and AUTO_CONFIRM_SAFE:
            confirm = "y"

        else:

            confirm = input(f"{YELLOW}Change price? (y/n): {RESET}").strip().lower()

        if confirm != "y":
            continue

        print(f"{BLUE}Updating:{RESET} " f"${web['price']} -> ${local['price']}")

        status, text = patch_price(
            session,
            web["id"],
            web["variant_id"],
            int(local["price"] * 1.07),
        )

        print(f"PATCH STATUS: {status}")

        if status != 200:
            print(text)

    print()

    print(f"{GREEN}Done.{RESET}")


# ============================================================

if __name__ == "__main__":
    main()
